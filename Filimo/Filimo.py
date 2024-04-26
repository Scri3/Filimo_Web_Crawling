# Importings
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import time

# Creating a driver
driver = webdriver.Chrome(executable_path=r"E:\Chrome Driver\chromedriver.exe")

driver.maximize_window()

action = ActionChains(driver)

# URL
driver.get("https://www.filimo.com/asparagus")

time.sleep(1)

# Scrolling page
def Scrolling_to_the_bottom():

    while True:

        # Scroll down 1000 pixels
        driver.execute_script('window.scrollBy(0, 2500)')

        # Wait for page to load
        time.sleep(1.5)

        # Check if at bottom of page
        if driver.execute_script('return window.innerHeight + window.pageYOffset >= document.body.offsetHeight'):
            break

Scrolling_to_the_bottom()


# Finding Elements
Categories = driver.find_elements(By.CSS_SELECTOR,".styles_row-header-links__6JyOl")

CategoryNames = driver.find_elements(By.CSS_SELECTOR,".styles_row-header-title__PWYdu")

NewFile = openpyxl.Workbook()
NewFile.save("Filimo.xlsx")


for (c,cn) in zip(Categories, CategoryNames):
    
    action.scroll_to_element(c).perform()

    CategoryName = cn.text

    NewLink = c.find_element(By.TAG_NAME, "a").get_attribute('href')

    driver.execute_script("window.open('');")

    time.sleep(1)


    driver.switch_to.window(driver.window_handles[1])

    driver.get(NewLink)

    time.sleep(2)

    Scrolling_to_the_bottom()
    
    MovieThumbnails = driver.find_elements(By.CSS_SELECTOR, ".styles_thumbnail-wrapper__FDMIk")

    Title = []

    EnglishTitle = []

    Director = []

    Country = []

    Year = []

    Genre1 = []

    Genre2 = []

    URL = []

    TED_List = [Title, EnglishTitle, Director]
    TED_Tags = [".styles_title__3tlZr", ".styles_title-en__6-yb9", ".styles_director-value__Ut\+gI"]
    
    for m in MovieThumbnails:

        action.scroll_to_element(m).perform()

        NewLink2 = m.find_element(By.TAG_NAME, "a").get_attribute('href')

        driver.execute_script("window.open('');")

        time.sleep(1)

        driver.switch_to.window(driver.window_handles[2])

        driver.get(NewLink2)

        time.sleep(2)

        

        for i,j in zip(TED_List, TED_Tags):
            try:
                i.append(driver.find_element(By.CSS_SELECTOR, j).text)

            except:
                i.append("")
                

        Desc = driver.find_elements(By.CSS_SELECTOR, ".styles_details-movie-description-text__uMwtT")
        LengDesc = len(Desc)

        if Desc[LengDesc - 1].text == "کیفیت HD":
            try:
                Country.append(Desc[LengDesc - 3].text)
                Year.append(Desc[LengDesc - 2].text)
            except:
                Country.append("")
                Year.append("")

        else:
            try:
                Country.append(Desc[LengDesc - 2].text)
                Year.append(Desc[LengDesc - 1].text)
            except:
                Country.append("")
                Year.append("")


        Genres = driver.find_elements(By.CSS_SELECTOR, ".styles_badge-natural-light-outline__34vOE")
        LengGenres = len(Genres)

        if LengGenres == 2:
            try:
                Genre1.append(Genres[0].text)
                Genre2.append(Genres[1].text)
            except:
                Genre1.append("")
                Genre2.append("")
                

        else:
            try:
                Genre1.append(Genres[0].text)
            except:
                Genre1.append("")

            Genre2.append("")

        get_url = driver.current_url

        URL.append(get_url)

        driver.close()

        driver.switch_to.window(driver.window_handles[1])

        time.sleep(2)

    # Putting the crawled data in dictionary
    Data = {'Title': Title, 'EnglishTitle': EnglishTitle,'Director': Director, 'Country': Country, 'Year': Year,
    'Genre 1': Genre1, 'Genre 2': Genre2, 'Link': URL}

    # Creating a frame
    df = pd.DataFrame(Data)

    # Exporting to excel
    datatoexcel = pd.ExcelWriter('Filimo.xlsx', engine='openpyxl', mode= 'a', if_sheet_exists='overlay')
    df.to_excel(datatoexcel, sheet_name = CategoryName, index= False)

    Data.clear()

    # save the excel
    datatoexcel.close()


    driver.close()

    driver.switch_to.window(driver.window_handles[0])


driver.close()


wb = load_workbook("Filimo.xlsx")

if 'Sheet' in NewFile.sheetnames:
    wb.remove(wb['Sheet'])
    wb.save("Filimo.xlsx")


print("Crawled Successfully!")